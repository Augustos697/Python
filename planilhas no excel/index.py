import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
#Como visualizar paginas existentes 
print(book.sheetnames)
book.create_sheet('Frutas')
#Como selecionar uma página
frutas_pages = book["Frutas"]
frutas_pages.append(['Fruta','Quantidade','Custo'])
frutas_pages.append(['banana', '5','R$3,90'])
frutas_pages.append(['Maça', '3','R$6,90'])
frutas_pages.append(['Morango', '10','R$7,00'])
#Salver a planilha
book.save("Planilha de compras.xlsx")